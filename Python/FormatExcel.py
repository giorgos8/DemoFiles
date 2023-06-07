
# pip install openpyxl
import os
import sys
import openpyxl
from openpyxl.styles import Font, Color, PatternFill
from pathlib import Path

print("Formatting Excel for Documentation!!!")
# print(os.environ.get('PYTHON_PROGRAM_PATH'))

PythonProgramPath = "C:\\Python\\My Python Programs\\" #os.environ.get('PYTHON_PROGRAM_PATH')

# get arguments
# print(sys.argv[0])
# print(sys.argv[1])

#ExcelName = sys.argv[1]
ExcelName = "DB_Doc"

print(ExcelName)

#FullPath = str("C:\\Python\\My Python Programs\\" + ExcelName + ".xlsx")

#FullPath = Path("C:/Python/My Python Programs/" + ExcelName + ".xlsx")
FullPath = Path(PythonProgramPath + ExcelName + ".xlsx")
print(FullPath)

#filename = Path("source_data/text_files/raw_data.txt")
#print(filename.name)

# "C:\\Python\\My Python Programs\\RiskAvert_DB_Doc.xlsx"
# wb = openpyxl.load_workbook("C:\\Python\\My Python Programs\\" + ExcelName + ".xlsx")
wb = openpyxl.load_workbook(FullPath)

# list with all Excel Sheets
# print(wb.sheetnames)

# Get a specific WorkSheet
ws_Index = wb['Index']

# Create a new Sheet on the WorkBook (at the end of the Sheets)
# wb.create_sheet('NewSheet')

# Create a new Sheet on the WorkBook (at the 0 position)
# wb.create_sheet('FirstSheet', 0)

# Get-Read Specific Cell of the specific Sheet
# print(ws_Index['C3'].value)

# Get-Read Specific Cell of the specific Sheet - another way
# print(ws_Index.cell(row=5, column=3).value)

# Get value range
# value_range = ws_Index['A1':'C7']

# for a, b, c in value_range:
#    print(a.value, b.value, c.value)

# Iterate Rows
# rows = ws_Index.iter_rows(min_row=1, max_row=7, min_col=1, max_col=3)

# for a, b, c in rows:
#    print(a.value, b.value, c.value)


# Iterate Columns
# columns = ws_Index.iter_cols(min_row=1, max_row=7, min_col=1, max_col=3)

# for col in columns:
#    print(col)

  
# rows
# rows = list(ws_Index.rows)
# print(rows)


# columns = list(ws_Index.columns)
# print(columns)

# setting fonts
font_style_blue = Font(color='3377FF')
fill_pattern = PatternFill(fgColor='C64747', patternType="solid")

# write/update data
ws_Index['A1'].value = "SHEMA_NAME"
#ws_Index['A1'].font = font_style_blue

ws_Index['B1'].value = "TABLE_NAME"
#ws_Index['B1'].font = font_style_blue

ws_Index['C1'].value = ""
ws_Index['D1'].value = "TABLE_DESCRIPTION"

# fill backround color
# ws_Index['B3'].fill = fill_pattern

for c in range(5, 20):
    ws_Index.cell(row = 1, column=c).value = ""    
    

for i in range(2, 1000):
    if ws_Index.cell(row = i, column=3).value is not None:
        ws_Index.cell(row = i, column=2).value = ws_Index.cell(row = i, column=3).value        
    ws_Index.cell(row = i, column=3).value = ""


# Iterate Sheet Names
s = 2
for sheet in wb.worksheets:    
    if sheet.title == "Index":
        continue
    #print(sheet.title)    
    link = "RiskAvert_DB_Doc.xlsx" + "#" + sheet.title
    ws_Index.cell(row=s, column=2).hyperlink = (link)
    ws_Index.cell(row=s, column=2).font = font_style_blue
    s = s + 1
    

# print(wb.sheetnames)

# Hyperlink to Excel
#link = "RiskAvert_DB_Doc.xlsx#RMF_LEVERAGE_TRANCHES"
#ws_Index.cell(row=2, column=2).hyperlink = (link)



# Save the Excel Workbook
# wb.save("C:\\Python\\My Python Programs\\RiskAvert_DB_Doc.xlsx");
# wb.save("C:\\Python\\My Python Programs\\" + ExcelName + ".xlsx");
wb.save(FullPath);





### from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
### font = Font(name='Calibri',
###                 size=11,
###                 bold=False,
###                 italic=False,
###                 vertAlign=None,
###                 underline='none',
###                 strike=False,
###                 color='FF000000')
### fill = PatternFill(fill_type=None,
###                 start_color='FFFFFFFF',
###                 end_color='FF000000')
### border = Border(left=Side(border_style=None,
###                           color='FF000000'),
###                 right=Side(border_style=None,
###                            color='FF000000'),
###                 top=Side(border_style=None,
###                          color='FF000000'),
###                 bottom=Side(border_style=None,
###                             color='FF000000'),
###                 diagonal=Side(border_style=None,
###                               color='FF000000'),
###                 diagonal_direction=0,
###                 outline=Side(border_style=None,
###                              color='FF000000'),
###                 vertical=Side(border_style=None,
###                               color='FF000000'),
###                 horizontal=Side(border_style=None,
###                                color='FF000000')
###                )
### alignment=Alignment(horizontal='general',
###                     vertical='bottom',
###                     text_rotation=0,
###                     wrap_text=False,
###                     shrink_to_fit=False,
###                     indent=0)
### number_format = 'General'
### protection = Protection(locked=True,
###                         hidden=False)
###
